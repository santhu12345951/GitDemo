import openpyxl


class HomePageData:


    test_HomePageData= [{"firstname":"Rahul","lastname":"shetty","Gender":"Male"},{"firstname":"santhosh","lastname":"Krishna","Gender":"Female"}]

    @staticmethod
    def gettestdata(testcase_name):


        book = openpyxl.load_workbook("C:\\Users\\Santhosh Krishna\\Contacts\\Documents\\PythonDemo.xlsx")

        sheet = book.active
        Dict= {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

